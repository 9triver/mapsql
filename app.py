#!/usr/bin/env python3
"""
MapSQL Web UI — Excel 映射定义校验与 SQL 生成。
用法: python app.py [--port 5000]
"""

import json
import os
import tempfile

from flask import Flask, jsonify, request, render_template, send_file

from generate_sql import ExcelParser, SQLGenerator

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# 存储上传文件的临时路径（进程生命周期内有效）
_uploaded_file: str | None = None


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload():
    """上传 Excel 文件，返回所有 sheet 名称列表。"""
    global _uploaded_file

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify(error='未选择文件'), 400

    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify(error='仅支持 .xlsx / .xls 文件'), 400

    # 保存到临时文件
    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix='.xlsx', prefix='mapsql_'
    )
    f.save(tmp.name)
    tmp.close()

    # 清理旧文件
    if _uploaded_file and os.path.exists(_uploaded_file):
        os.unlink(_uploaded_file)
    _uploaded_file = tmp.name

    # 读取 sheet 列表
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return jsonify(error=f'无法读取 Excel: {e}'), 400

    return jsonify(filename=f.filename, sheets=sheets)


@app.route('/api/sheet/<sheet_name>')
def get_sheet_data(sheet_name):
    """返回指定 sheet 的全部单元格数据（用于前端表格渲染）。"""
    if not _uploaded_file:
        return jsonify(error='请先上传文件'), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(_uploaded_file, data_only=True)
        if sheet_name not in wb.sheetnames:
            return jsonify(error=f"Sheet '{sheet_name}' 不存在"), 404

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            max_col=ws.max_column, values_only=True
        ):
            rows.append([
                str(cell) if cell is not None else ''
                for cell in row
            ])
        wb.close()

        # 找到实际有内容的最大列数（避免大量空列）
        max_col = 0
        for row in rows:
            for i in range(len(row) - 1, -1, -1):
                if row[i]:
                    max_col = max(max_col, i + 1)
                    break
        # 截断空列
        rows = [row[:max_col] for row in rows]

    except Exception as e:
        return jsonify(error=f'读取失败: {e}'), 500

    return jsonify(rows=rows, total_rows=len(rows), total_cols=max_col)


@app.route('/api/validate/<sheet_name>')
def validate_sheet(sheet_name):
    """校验 sheet 是否符合映射定义规则。"""
    if not _uploaded_file:
        return jsonify(error='请先上传文件'), 400

    ep = ExcelParser(_uploaded_file, sheet_name)
    mapping = ep.parse()

    result = {
        'valid': mapping is not None,
        'warnings': ep.warnings,
        'errors': ep.errors,
        'summary': None,
    }

    if mapping:
        segments_info = []
        for seg in mapping.segments:
            segments_info.append({
                'name': seg.segment_name,
                'source_tables': len(seg.source_tables),
                'where_conditions': len(seg.where_conditions),
                'field_mappings': len(seg.field_mappings),
            })
        result['summary'] = {
            'target_table': mapping.target_table,
            'target_cn_name': mapping.target_cn_name,
            'segments': segments_info,
        }

    return jsonify(result)


@app.route('/api/generate/<sheet_name>')
def generate_sql(sheet_name):
    """生成 SQL 存储过程。"""
    if not _uploaded_file:
        return jsonify(error='请先上传文件'), 400

    ep = ExcelParser(_uploaded_file, sheet_name)
    mapping = ep.parse()

    if not mapping:
        return jsonify(
            error='解析失败',
            errors=ep.errors,
            warnings=ep.warnings,
        ), 400

    gen = SQLGenerator(mapping)
    sql = gen.generate()

    return jsonify(
        sql=sql,
        notes=gen.notes,
        warnings=ep.warnings,
        target_table=mapping.target_table,
    )


@app.route('/api/download/<sheet_name>')
def download_sql(sheet_name):
    """下载生成的 SQL 文件。"""
    if not _uploaded_file:
        return jsonify(error='请先上传文件'), 400

    ep = ExcelParser(_uploaded_file, sheet_name)
    mapping = ep.parse()

    if not mapping:
        return jsonify(error='解析失败'), 400

    gen = SQLGenerator(mapping)
    sql = gen.generate()

    # 写入临时文件供下载
    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix='.sql', prefix='mapsql_',
        mode='w', encoding='utf-8'
    )
    tmp.write(sql)
    tmp.close()

    filename = f'{mapping.target_table}.sql'
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype='text/plain',
    )


if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser(description='MapSQL Web UI')
    p.add_argument('--port', type=int, default=6000)
    p.add_argument('--host', default='127.0.0.1')
    args = p.parse_args()

    print(f'MapSQL Web UI: http://{args.host}:{args.port}')
    app.run(host=args.host, port=args.port, debug=True)
