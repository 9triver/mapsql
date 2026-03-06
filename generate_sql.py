#!/usr/bin/env python3
"""
根据 Excel 映射定义生成 MySQL 存储过程 SQL。
用法: python generate_sql.py <excel_file> <sheet_name> [-o output_file]
"""

import argparse
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 数据结构
# ============================================================

@dataclass
class SourceTable:
    """数据源表定义"""
    table_name: str          # 表英文名
    table_cn_name: str       # 表中文名
    alias: str               # 别名 (T1, T2, ...)
    join_type: str           # 关联类型: '' (主表) / 'LEFT JOIN' / 'INNER JOIN'
    join_condition: str      # 关联条件 (ON ...)
    remark: str = ''         # 备注


@dataclass
class WhereCondition:
    """数据范围条件"""
    operator: str            # WHERE / AND
    condition: str           # 条件表达式
    description: str = ''    # 逻辑说明


@dataclass
class FieldMapping:
    """字段映射"""
    target_cn_name: str      # 目标字段中文名
    target_en_name: str      # 目标字段英文名
    target_type: str         # 目标字段类型
    target_dict: str         # 字典枚举
    source_table: str        # 源表名
    source_field: str        # 源字段英文名
    source_cn_name: str      # 源字段中文名
    source_type: str         # 源字段类型
    mapping_rule: str        # 映射规则
    source_dict: str         # 源系统字典
    fill_instruction: str    # 填报说明（Col12）
    description: str         # 业务口径（Col13）
    biz_scope: str           # 业务范围（Col14）


@dataclass
class MappingSegment:
    """一段映射块（A段/B段等）"""
    segment_name: str                          # 段名称
    source_tables: list = field(default_factory=list)   # 数据源表列表
    where_conditions: list = field(default_factory=list) # WHERE 条件列表
    field_mappings: list = field(default_factory=list)   # 字段映射列表
    alias_map: dict = field(default_factory=dict)  # 旧别名→新别名映射


@dataclass
class SheetMapping:
    """整个 Sheet 的映射定义"""
    target_table: str        # 目标表英文名
    target_cn_name: str      # 目标表中文名
    segments: list = field(default_factory=list)  # 映射段列表


# ============================================================
# Excel 解析器
# ============================================================

class ExcelParser:
    """解析 Excel Sheet 中的映射定义"""

    # 全角→半角映射表
    _FULLWIDTH_MAP = str.maketrans(
        '（），；＝＋', '(),;=+',
    )

    def __init__(self, filepath: str, sheet_name: str):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.warnings: list[str] = []
        self.errors: list[str] = []

    def _warn(self, row: int, msg: str):
        self.warnings.append(f"[警告] 行{row}: {msg}")

    def _error(self, row: int, msg: str):
        self.errors.append(f"[错误] 行{row}: {msg}")

    @classmethod
    def _clean_sql_text(cls, text: str) -> str:
        """清洗 Excel 文本中的非标准字符，使其成为合法 SQL 片段。
        - 全角字符转半角
        - 修复 ON 条件中的拼接错误（如 IDAND → ID AND）
        - 清理特殊字符干扰（如 DATE+ID → DATE_ID）
        """
        if not text:
            return text
        # 全角→半角
        text = text.translate(cls._FULLWIDTH_MAP)
        # 修复关键字与标识符粘连：在标识符与SQL关键字之间插入空格
        # 如 GUAR_CONTRACT_IDAND → GUAR_CONTRACT_ID AND
        text = re.sub(
            r'([A-Z0-9_])(AND|OR|ON|LEFT|RIGHT|INNER|JOIN|WHERE)\b',
            r'\1 \2', text, flags=re.IGNORECASE
        )
        # 修复 T3.=DATE_ID → T3.DATE_ID
        text = re.sub(r'(\w)\.=(\w)', r'\1.\2', text)
        return text

    def _cell(self, row, col) -> str:
        """安全读取单元格值，返回去空格字符串"""
        val = self.ws.cell(row=row, column=col).value
        if val is None:
            return ''
        return str(val).strip()

    def parse(self) -> Optional[SheetMapping]:
        """解析整个 Sheet"""
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
        except FileNotFoundError:
            self.errors.append(f"[错误] 文件不存在: {self.filepath}")
            return None
        except Exception as e:
            self.errors.append(f"[错误] 无法打开文件: {e}")
            return None

        if self.sheet_name not in wb.sheetnames:
            self.errors.append(
                f"[错误] Sheet '{self.sheet_name}' 不存在。"
                f"可用的 Sheet: {', '.join(wb.sheetnames)}"
            )
            return None

        self.ws = wb[self.sheet_name]
        max_row = self.ws.max_row
        max_col = self.ws.max_column

        # 第1步：找目标表定义（第2行）
        target_table, target_cn_name = self._parse_target_table()
        if not target_table:
            return None

        mapping = SheetMapping(
            target_table=target_table,
            target_cn_name=target_cn_name,
        )

        # 第2步：扫描所有行，按段落分割
        segments_raw = self._split_segments(max_row, max_col)

        for seg_name, seg_rows in segments_raw:
            segment = self._parse_segment(seg_name, seg_rows)
            if segment:
                mapping.segments.append(segment)

        if not mapping.segments:
            self.errors.append("[错误] 未找到任何有效的映射段")
            return None

        return mapping

    def _parse_target_table(self) -> tuple[str, str]:
        """解析目标表名（通常在第2行）"""
        for r in range(1, min(self.ws.max_row + 1, 5)):
            if self._cell(r, 1) == '目标表：':
                table_name = self._cell(r, 2)
                table_cn = self._cell(r, 3)
                if not table_name:
                    self._error(r, "目标表名为空")
                    return '', ''
                return table_name, table_cn
        self._error(0, "未找到 '目标表：' 定义行。预期格式：第1列为'目标表：'，第2列为表英文名，第3列为表中文名")
        return '', ''

    def _split_segments(self, max_row: int, max_col: int) -> list:
        """将 Sheet 按段落分割。检测 '数据源表：' 出现的位置作为段落起点"""
        segment_starts = []
        for r in range(1, max_row + 1):
            c1 = self._cell(r, 1)
            # 检测段名称行（如 "A段：表内贷款借据"）
            if re.match(r'^[A-Z]段[：:]', c1):
                seg_name = c1
                continue
            # 检测数据源表起始行
            if c1 == '数据源表：':
                name = seg_name if 'seg_name' in dir() and seg_name else '默认段'
                segment_starts.append((name, r))
                seg_name = ''  # reset

        # 如果没找到段名称标记，但找到了数据源表
        if not segment_starts:
            for r in range(1, max_row + 1):
                if self._cell(r, 1) == '数据源表：':
                    segment_starts.append(('默认段', r))
                    break

        if not segment_starts:
            self._error(0, "未找到 '数据源表：' 定义行")
            return []

        # 确定每段的行范围
        result = []
        for i, (name, start_row) in enumerate(segment_starts):
            if i + 1 < len(segment_starts):
                end_row = segment_starts[i + 1][1] - 1
            else:
                end_row = max_row
            result.append((name, (start_row, end_row)))

        return result

    def _parse_segment(self, seg_name: str, row_range: tuple) -> Optional[MappingSegment]:
        """解析一个映射段"""
        start_row, end_row = row_range
        segment = MappingSegment(segment_name=seg_name)

        # 在段内定位三个区域
        source_table_start = start_row
        where_start = None
        field_start = None

        for r in range(start_row, end_row + 1):
            c1 = self._cell(r, 1)
            if c1 == '数据范围条件：':
                where_start = r
            if c1 == '字段映射':
                field_start = r

        # 解析数据源表
        source_end = (where_start - 1) if where_start else (
            field_start - 1 if field_start else end_row)
        segment.source_tables = self._parse_source_tables(
            source_table_start, source_end)
        # 提取别名映射（由 _parse_source_tables 生成）
        segment.alias_map = getattr(self, '_last_alias_map', {})

        if not segment.source_tables:
            self._error(start_row, f"段 '{seg_name}' 未找到任何数据源表定义")
            return None

        # 解析 WHERE 条件
        if where_start:
            where_end = (field_start - 1) if field_start else end_row
            segment.where_conditions = self._parse_where_conditions(where_start, where_end)
            # 应用别名映射到 WHERE 条件
            if segment.alias_map:
                for wc in segment.where_conditions:
                    wc.condition = self._replace_aliases(
                        wc.condition, segment.alias_map)
            # 对关联类型为"无"的表，将 WHERE 中的别名引用展开为 "表名 别名"
            for t in segment.source_tables:
                if t.join_type == '' and t != segment.source_tables[0]:
                    wc_alias = t.alias
                    for wc in segment.where_conditions:
                        # From T2 → From table_name T2
                        wc.condition = re.sub(
                            rf'\bFrom\s+{re.escape(wc_alias)}\b',
                            f'From {t.table_name} {wc_alias}',
                            wc.condition,
                            flags=re.IGNORECASE,
                        )

        # 解析字段映射
        if field_start:
            segment.field_mappings = self._parse_field_mappings(field_start, end_row)
        else:
            self._error(start_row, f"段 '{seg_name}' 未找到 '字段映射' 标记行")
            return None

        if not segment.field_mappings:
            self._error(start_row, f"段 '{seg_name}' 未解析到任何字段映射")
            return None

        return segment

    def _parse_source_tables(self, start_row: int, end_row: int) -> list:
        """解析数据源表区"""
        tables = []
        # 标题行在 start_row，数据从 start_row+1 开始
        for r in range(start_row + 1, end_row + 1):
            table_name = self._cell(r, 2)
            if not table_name:
                continue
            table_cn = self._cell(r, 3)
            alias_raw = self._cell(r, 4)
            join_type_raw = self._cell(r, 5)
            join_cond_raw = self._cell(r, 6)
            remark = self._cell(r, 8)

            # 提取纯别名：从 "主表 A" / "T3 法定代表人" 中提取 SQL 标识符
            alias = self._extract_alias(alias_raw, len(tables) + 1)
            if not alias:
                self._warn(r, f"表 {table_name} 缺少别名，自动分配 T{len(tables) + 1}")
                alias = f'T{len(tables) + 1}'

            # 标准化关联类型
            join_type = ''
            if join_type_raw:
                jt = join_type_raw.upper().strip()
                if jt == '无' or jt == 'NONE':
                    join_type = ''  # 不生成 JOIN，仅在 WHERE 子查询中引用
                elif 'INNER' in jt:
                    join_type = 'INNER JOIN'
                elif 'LEFT' in jt:
                    join_type = 'LEFT JOIN'
                elif 'RIGHT' in jt:
                    join_type = 'RIGHT JOIN'
                else:
                    self._warn(r, f"无法识别关联类型 '{join_type_raw}'，默认为 LEFT JOIN")
                    join_type = 'LEFT JOIN'

            # 清理关联条件（去掉开头的 ON，清洗全角/拼写错误）
            join_cond = self._clean_sql_text(join_cond_raw)
            if join_cond.upper().startswith('ON '):
                join_cond = join_cond[3:].strip()

            tables.append(SourceTable(
                table_name=table_name,
                table_cn_name=table_cn,
                alias=alias,
                join_type=join_type,
                join_condition=join_cond,
                remark=remark,
            ))

        if tables and tables[0].join_type:
            self._warn(start_row + 1, f"第一个表 {tables[0].table_name} 不应有关联类型（应为主表）")

        # 统一别名为 Tn 格式：将单字母别名 (A,B,C,...)
        # 替换为 T1,T2,T3,...
        alias_map = {}  # old_alias → new_alias
        for i, t in enumerate(tables):
            new_alias = f'T{i + 1}'
            if t.alias != new_alias and re.match(r'^[A-Z]$', t.alias):
                alias_map[t.alias] = new_alias
                t.alias = new_alias

        # 替换 JOIN 条件中的旧别名
        if alias_map:
            for t in tables:
                t.join_condition = self._replace_aliases(
                    t.join_condition, alias_map
                )

        # 保存别名映射供后续使用
        self._last_alias_map = alias_map
        return tables

    @staticmethod
    def _replace_aliases(text: str, alias_map: dict) -> str:
        """替换 SQL 文本中的表别名引用。
        如将 A.FIELD → T1.FIELD, From C Where → From T2 Where"""
        if not text or not alias_map:
            return text
        for old, new in alias_map.items():
            # 替换 A.FIELD → T1.FIELD
            text = re.sub(
                rf'\b{re.escape(old)}\.',
                f'{new}.', text
            )
            # 替换独立别名引用（如 From C Where → From T2 Where）
            text = re.sub(
                rf'\b{re.escape(old)}\b(?!\.)',
                new, text
            )
        return text

    @staticmethod
    def _extract_alias(alias_raw: str, default_idx: int) -> str:
        """从别名字段中提取 SQL 别名。
        如 '主表 A' → 'A', 'T3 法定代表人' → 'T3',
           '主表A' → 'A', '关联表B' → 'B'"""
        if not alias_raw:
            return ''
        parts = alias_raw.split()
        # 查找 T\d+ 模式（空格分隔或尾部）
        for p in parts:
            if re.match(r'^T\d+$', p, re.IGNORECASE):
                return p
        # 在整个字符串中查找 T\d+ 模式（如 "关联表T3"）
        m = re.search(r'(T\d+)', alias_raw, re.IGNORECASE)
        if m:
            return m.group(1)
        # 查找单字母别名 (A, B, etc.)
        for p in parts:
            if re.match(r'^[A-Z]$', p):
                return p
        # 查找末尾单字母（如 "主表A", "关联表B"）
        m = re.search(r'([A-Z])$', alias_raw)
        if m:
            return m.group(1)
        # fallback: 如果第一个 token 是纯英文
        if parts and re.match(r'^[a-zA-Z_]\w*$', parts[0]):
            return parts[0]
        # 如果只有中文（如 "主表"），返回空
        return ''

    def _parse_where_conditions(self, start_row: int, end_row: int) -> list:
        """解析数据范围条件区"""
        conditions = []
        for r in range(start_row + 1, end_row + 1):
            operator = self._cell(r, 2).upper()
            condition = self._clean_sql_text(self._cell(r, 3))
            description = self._cell(r, 5)
            if not operator or not condition:
                continue
            # 转换中文伪函数：月初(X) → DATE_FORMAT(X, '%Y-%m-01')
            condition = re.sub(
                r'月初\((\w+)\)',
                r"DATE_FORMAT(\1, '%Y-%m-01')",
                condition,
            )
            # 转换 Oracle 语法
            condition = self._convert_oracle_syntax(condition)
            conditions.append(WhereCondition(
                operator=operator,
                condition=condition,
                description=description,
            ))
        return conditions

    @staticmethod
    def _convert_oracle_syntax(text: str) -> str:
        """将 Oracle 语法转换为 MySQL 语法（用于 WHERE/JOIN 条件）"""
        if not text:
            return text
        # NVL(a, b) → IFNULL(a, b)
        text = re.sub(r'\bNVL\s*\(', 'IFNULL(', text, flags=re.IGNORECASE)
        # TO_DATE('...', '...') → STR_TO_DATE('...', '%Y-%m-%d')
        text = re.sub(
            r"\bTO_DATE\s*\(\s*'([^']+)'\s*,\s*'[^']+'\s*\)",
            r"STR_TO_DATE('\1', '%Y-%m-%d')",
            text, flags=re.IGNORECASE,
        )
        # TO_CHAR(field, 'YYYY-MM-DD') → DATE_FORMAT(field, '%Y-%m-%d')
        text = re.sub(
            r"\bTO_CHAR\s*\(([^,]+),\s*'YYYY-MM-DD'\s*\)",
            r"DATE_FORMAT(\1, '%Y-%m-%d')",
            text, flags=re.IGNORECASE,
        )
        return text

    def _parse_field_mappings(self, start_row: int, end_row: int) -> list:
        """解析字段映射区"""
        mappings = []

        # 跳过标题行：找到 '字段中文名' 所在行
        header_row = None
        for r in range(start_row, min(start_row + 5, end_row + 1)):
            if self._cell(r, 1) == '字段中文名':
                header_row = r
                break

        if header_row is None:
            # 也尝试查找 '目标字段' 行后面的行
            for r in range(start_row, min(start_row + 5, end_row + 1)):
                if '目标字段' in self._cell(r, 1) or '字段中文名' in self._cell(r, 1):
                    header_row = r
                    break

        if header_row is None:
            self._error(start_row, "字段映射区未找到标题行（'字段中文名'）")
            return []

        # 从标题行下一行开始解析
        for r in range(header_row + 1, end_row + 1):
            cn_name = self._cell(r, 1)
            en_name = self._cell(r, 2)

            # 跳过空行
            if not cn_name and not en_name:
                continue

            # 跳过辅助行（同一字段的多个源字段，如贷款逾期标识的 INT_OVERDUE_BAL）
            if not en_name and not self._cell(r, 3):
                # 这是一个补充行，源字段信息附加到上一个映射
                src_field = self._cell(r, 7)
                if src_field and mappings:
                    mappings[-1].mapping_rule = mappings[-1].mapping_rule or ''
                    # 标记有多源字段
                    if not mappings[-1].mapping_rule:
                        mappings[-1].mapping_rule = f'__MULTI_SRC__:{src_field}'
                continue

            target_type = self._cell(r, 3)
            target_dict = self._cell(r, 4)
            source_table = self._cell(r, 6)
            source_field = self._cell(r, 7)
            source_cn = self._cell(r, 8)
            source_type = self._cell(r, 9)
            mapping_rule = self._cell(r, 10)
            source_dict = self._cell(r, 11)
            fill_instruction = self._cell(r, 12)
            description = self._cell(r, 13)
            biz_scope = self._cell(r, 14)

            # 校验
            if en_name and not target_type and cn_name != '采集日期':
                self._warn(r, f"字段 {en_name}({cn_name}) 缺少目标字段类型")

            mappings.append(FieldMapping(
                target_cn_name=cn_name,
                target_en_name=en_name,
                target_type=target_type,
                target_dict=target_dict,
                source_table=source_table,
                source_field=source_field,
                source_cn_name=source_cn,
                source_type=source_type,
                mapping_rule=mapping_rule,
                source_dict=source_dict,
                fill_instruction=fill_instruction,
                description=description,
                biz_scope=biz_scope,
            ))

        return mappings


# ============================================================
# CASE 字典提取器
# ============================================================

class CaseDictExtractor:
    """从手写 SQL 文件中提取 CASE WHEN 映射字典。

    字典键为 (目标字段, 源字段)，值为 CASE 表达式字符串。
    同一源字段在不同目标字段中可能有不同映射（如 CUST_TYPE_CODE）。
    """

    def __init__(self):
        self.case_dict: dict[tuple[str, str], str] = {}

    @staticmethod
    def _parse_select_exprs(select_text: str) -> list[str]:
        """按顶层逗号拆分 SELECT 表达式，正确处理 CASE/括号嵌套。"""
        # 先去除行尾 SQL 注释（保留字符串内的 --）
        lines = select_text.split('\n')
        clean_lines = []
        for line in lines:
            # 简单处理：去掉 -- 后的内容（不在引号内）
            in_str = False
            result = []
            i = 0
            while i < len(line):
                if line[i] == "'" and not in_str:
                    in_str = True
                elif line[i] == "'" and in_str:
                    in_str = False
                elif line[i:i+2] == '--' and not in_str:
                    break
                result.append(line[i])
                i += 1
            clean_lines.append(''.join(result).rstrip())
        cleaned = '\n'.join(clean_lines)

        tokens = re.findall(r"'[^']*'|\b\w+\b|[(),]", cleaned)
        exprs: list[str] = []
        paren_depth = 0
        case_depth = 0
        last_split = 0
        pos = 0

        for token in tokens:
            tup = token.upper()
            if tup == '(':
                paren_depth += 1
            elif tup == ')':
                paren_depth -= 1
            elif tup == 'CASE':
                case_depth += 1
            elif tup == 'END':
                case_depth = max(0, case_depth - 1)
            elif token == ',' and paren_depth <= 0 and case_depth <= 0:
                idx = cleaned.find(',', pos)
                if idx >= 0:
                    expr = cleaned[last_split:idx].strip()
                    if expr:
                        exprs.append(expr)
                    last_split = idx + 1
                    pos = idx + 1
                    continue

            idx = cleaned.find(token, pos)
            if idx >= 0:
                pos = idx + len(token)

        last = cleaned[last_split:].strip()
        if last:
            exprs.append(last)
        return exprs

    def load_from_directory(self, dir_path: str) -> int:
        """从目录中的所有 .txt/.sql 文件提取 CASE 映射。返回提取到的条目数。"""
        import os
        count = 0
        for fname in sorted(os.listdir(dir_path)):
            if not fname.endswith(('.txt', '.sql')):
                continue
            filepath = os.path.join(dir_path, fname)
            count += self._extract_from_file(filepath)
        return count

    def _extract_from_file(self, filepath: str) -> int:
        """从单个 SQL 文件中提取 CASE 映射。"""
        with open(filepath, encoding='utf-8') as f:
            text = f.read()

        count = 0
        # 匹配 INSERT [/*+ hint */] INTO table (cols) \n select ... \nfrom
        # 使用 \nfrom (行首 from) 避免匹配 CASE 表达式内的 from
        for m in re.finditer(
            r'insert\s+(?:/\*.*?\*/\s*)?into\s+[\w.]+\s*\(([^)]+)\)\s*\n?\s*select\s+(.+?)\n\s*from\s',
            text, re.DOTALL | re.IGNORECASE,
        ):
            # 去除注释后再拆分列名
            raw_cols = m.group(1)
            raw_cols = re.sub(r'--[^\n]*', '', raw_cols)
            cols = [c.strip() for c in raw_cols.split(',') if c.strip()]
            exprs = self._parse_select_exprs(m.group(2))

            for j in range(min(len(cols), len(exprs))):
                expr = exprs[j].strip()
                # 去除行尾注释
                expr_clean = re.sub(
                    r'\s*--[^\n]*$', '', expr, flags=re.MULTILINE
                ).strip()
                if not re.match(r'CASE\b', expr_clean, re.IGNORECASE):
                    continue
                col = cols[j].strip()
                # 提取源字段：优先匹配 T\d+.FIELD 模式
                src_m = re.search(r'\bT\d+\.(\w+)', expr_clean)
                if not src_m:
                    continue
                src_field = src_m.group(1)
                # 规范化：去除表别名前缀（CASE 中可能是 T1/T4 等不同别名）
                key = (col, src_field)
                if key not in self.case_dict:
                    self.case_dict[key] = expr_clean
                    count += 1
        return count

    def lookup(self, target_field: str, source_field: str) -> Optional[str]:
        """查找 CASE 映射。返回 CASE 表达式或 None。"""
        return self.case_dict.get((target_field, source_field))


# ============================================================
# SQL 生成器
# ============================================================

class SQLGenerator:
    """根据解析结果生成 MySQL 存储过程"""

    def __init__(self, mapping: SheetMapping,
                 case_dict: Optional[CaseDictExtractor] = None):
        self.mapping = mapping
        self.case_dict = case_dict
        self.notes: list[str] = []  # 生成过程中的注意事项

    def _note(self, msg: str):
        self.notes.append(f"[注意] {msg}")

    def generate(self) -> str:
        """生成完整的存储过程 SQL"""
        m = self.mapping

        # 提取采集日期字段名（最后一个字段）
        date_field = self._find_date_field(m.segments[0])

        lines = []
        lines.append(self._gen_header(m, date_field))

        for i, seg in enumerate(m.segments, 1):
            lines.append(self._gen_segment(seg, i, len(m.segments)))

        lines.append(self._gen_footer())

        return '\n'.join(lines)

    def _find_date_field(self, segment: MappingSegment) -> str:
        """找到采集日期字段名"""
        for fm in reversed(segment.field_mappings):
            if '采集日期' in fm.target_cn_name or '采集' in fm.target_cn_name:
                return fm.target_en_name
            if fm.source_field == 'V_DATE' or fm.mapping_rule == 'V_DATE':
                return fm.target_en_name
        # fallback: 最后一个字段
        if segment.field_mappings:
            return segment.field_mappings[-1].target_en_name
        return 'DATE_FIELD'

    def _gen_header(self, m: SheetMapping, date_field: str) -> str:
        """生成存储过程头部"""
        # 收集所有源表信息用于注释
        all_tables = []
        for seg in m.segments:
            for t in seg.source_tables:
                info = f"--           {t.table_name} ({t.alias} {t.table_cn_name})"
                if info not in all_tables:
                    all_tables.append(info)

        tables_comment = '\n'.join(all_tables)
        if all_tables:
            tables_comment = all_tables[0].replace('--           ', '-- 源    表: ') + '\n' + '\n'.join(all_tables[1:])

        return f"""CREATE PROCEDURE Pids_{m.target_table.lower()}(
    IN I_DATE VARCHAR(8),   -- 数据日期，格式 YYYYMMDD
    OUT O_RLT VARCHAR(10)   -- 返回结果
)
BEGIN
    -- -------------------------------------------------------------------
    -- 功能描述: {m.target_cn_name} ({m.target_table})
    -- 传入参数: I_DATE 格式 YYYYMMDD
    {tables_comment}
    -- 目 标 表: {m.target_table}
    -- -------------------------------------------------------------------

    DECLARE V_DATE VARCHAR(8);
    DECLARE V_START_DT DATETIME;
    DECLARE V_PRD_NAME VARCHAR(100);
    DECLARE V_TAB_NAME VARCHAR(100);
    DECLARE V_TOTAL_NUM INT DEFAULT 0;
    DECLARE V_TAGS VARCHAR(300);
    DECLARE V_MSG TEXT;

    DECLARE EXIT HANDLER FOR SQLEXCEPTION
    BEGIN
        ROLLBACK;
        GET DIAGNOSTICS CONDITION 1 V_MSG = MESSAGE_TEXT;
        SET O_RLT = 'false';
        SET V_TAGS = CONCAT('第', V_TAGS, '段报错：', SUBSTRING(V_MSG, 1, 200));
        CALL dwdevdb_model.PMODEL_JOB_LOG(V_DATE, V_PRD_NAME, V_TAB_NAME, V_TAGS,
                          O_RLT, V_TOTAL_NUM, V_START_DT, NOW());
    END;

    -- 初始化变量
    SET V_DATE = I_DATE;
    SET V_START_DT = NOW();
    SET V_PRD_NAME = 'Pids_{m.target_table.lower()}';
    SET V_TAB_NAME = '{m.target_table}';
    SET V_TOTAL_NUM = 0;

    -- 删除当期数据（按采集日期）
    DELETE FROM {m.target_table} WHERE {date_field} = V_DATE;
"""

    # 聚合函数模式（用于检测 SELECT 表达式中的聚合）
    _AGG_FUNCS_RE = re.compile(
        r'\b(SUM|MAX|MIN|COUNT|AVG|GROUP_CONCAT)\s*\(',
        re.IGNORECASE
    )

    def _gen_segment(self, seg: MappingSegment,
                     seg_idx: int, total_segs: int) -> str:
        """生成一个映射段的 INSERT ... SELECT"""
        lines = []
        seg_label = (seg.segment_name
                     if seg.segment_name != '默认段'
                     else f'第{seg_idx}段')

        lines.append(
            f"    -- ======================== "
            f"{seg_label} BEGIN "
            f"========================")
        lines.append(f"    SET V_TAGS = '{seg_idx}';")
        lines.append("")

        # INSERT 字段列表
        valid_fields = [
            fm for fm in seg.field_mappings if fm.target_en_name]

        lines.append(
            f"    INSERT INTO {self.mapping.target_table} (")
        for i, fm in enumerate(valid_fields):
            comma = ',' if i < len(valid_fields) - 1 else ''
            lines.append(
                f"        {fm.target_en_name}{comma}"
                f"   -- {fm.target_cn_name}")
        lines.append("    )")

        # SELECT 字段列表 + 检测聚合函数
        lines.append("    SELECT")
        select_exprs = []
        has_aggregate = False
        for fm in valid_fields:
            expr = self._gen_select_expr(fm, seg)
            select_exprs.append(expr)
            if self._AGG_FUNCS_RE.search(expr):
                has_aggregate = True

        for i, (fm, expr) in enumerate(
            zip(valid_fields, select_exprs)
        ):
            comma = ',' if i < len(valid_fields) - 1 else ''
            lines.append(
                f"        {expr}{comma}"
                f"   -- {fm.target_cn_name}")

        # FROM / JOIN
        lines.append("")
        main_table = seg.source_tables[0]
        lines.append(
            f"    FROM {main_table.table_name}"
            f" {main_table.alias}")
        for t in seg.source_tables[1:]:
            if not t.join_type:
                continue  # 关联类型为"无"，不生成 JOIN（仅在 WHERE 子查询中引用）
            lines.append(
                f"    {t.join_type} {t.table_name} {t.alias}")
            lines.append(f"        ON {t.join_condition}")

        # WHERE
        if seg.where_conditions:
            for j, wc in enumerate(seg.where_conditions):
                prefix = "WHERE" if j == 0 else "  AND"
                lines.append(f"    {prefix} {wc.condition}")

        # GROUP BY（当 SELECT 中有聚合函数时）
        if has_aggregate:
            group_cols = []
            for fm, expr in zip(valid_fields, select_exprs):
                if not self._AGG_FUNCS_RE.search(expr):
                    # 排除常量表达式（空字符串、V_DATE、数字等）
                    stripped = expr.strip()
                    if (stripped.startswith("'") or
                            stripped == 'V_DATE' or
                            stripped.replace('.', '').isdigit()):
                        continue
                    group_cols.append(expr)
            if group_cols:
                lines.append("    GROUP BY")
                for i, col in enumerate(group_cols):
                    comma = ',' if i < len(group_cols) - 1 else ''
                    lines.append(f"        {col}{comma}")

        lines.append("    ;")

        lines.append("")
        lines.append(
            "    SET V_TOTAL_NUM = V_TOTAL_NUM + ROW_COUNT();")
        lines.append("    COMMIT;")
        lines.append(
            f"    -- ======================== "
            f"{seg_label} END "
            f"==========================")
        lines.append("")

        return '\n'.join(lines)

    def _gen_select_expr(self, fm: FieldMapping, seg: MappingSegment) -> str:
        """为一个字段映射生成 SELECT 表达式"""

        # 1. 采集日期 / V_DATE 参数 / 无源字段
        if fm.source_field == 'V_DATE':
            return 'V_DATE'
        if not fm.source_field and not fm.source_table:
            if '采集日期' in fm.target_cn_name or '采集' in fm.target_cn_name:
                return 'V_DATE'
            # 无源字段且非日期 → 空字符串（不能用 V_DATE）
            return "''"
        # 源字段为空但有源表名 → 空字符串
        if not fm.source_field:
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name}) "
                f"源表为 '{fm.source_table}' 但源字段为空，输出空字符串"
            )
            return "''"

        # 2. 源字段为中文名 → 不是合法列名
        if fm.source_field and re.search(
            r'[\u4e00-\u9fff]', fm.source_field
        ):
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name})"
                f" 源字段为中文 '{fm.source_field}'，"
                f"不是合法 SQL 列名，输出空字符串"
            )
            return "''"

        # 3. 源字段包含函数表达式 (NVL/IFNULL/COALESCE等)
        if fm.source_field and self._is_function_expr(fm.source_field):
            return self._convert_source_field_expr(fm, seg)

        # 3. 有显式映射规则
        if fm.mapping_rule and not fm.mapping_rule.startswith('__MULTI_SRC__'):
            return self._convert_mapping_rule(fm, seg)

        # 4. 多源字段标记
        if fm.mapping_rule and fm.mapping_rule.startswith('__MULTI_SRC__'):
            extra_field = fm.mapping_rule.split(':')[1]
            return self._gen_multi_source_expr(fm, extra_field, seg)

        # 5. 填报说明中有条件逻辑 ("当...时填...")
        if fm.fill_instruction and '当' in fm.fill_instruction:
            expr = self._gen_conditional_fill(fm, seg)
            if expr:
                return expr

        # 5b. 填报说明标注"需转换"
        if fm.fill_instruction and fm.fill_instruction.strip() in (
            '转换', '需转换'
        ):
            # 优先从字典查找 CASE 映射
            dict_expr = self._lookup_case_dict(fm, seg)
            if dict_expr:
                return dict_expr
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name}) "
                f"标注'需转换'，源字典值未知，暂直取"
            )

        # 5c. 源表字段包含逗号（多源表）→ 需人工处理
        if ',' in fm.source_table:
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name}) "
                f"引用多个源表 '{fm.source_table}'，"
                f"需人工补充取值逻辑"
            )

        # 6. 确定表别名
        alias = self._resolve_alias(fm, seg)

        # 7. 日期类型转换: 源为 DATE → 目标为 VARCHAR
        if self._is_date_conversion(fm):
            default = self._get_date_default(fm)
            expr = (f"DATE_FORMAT({alias}.{fm.source_field},"
                    f" '%Y-%m-%d')")
            if default:
                expr = f"COALESCE({expr}, '{default}')"
            return expr

        # 8. Y/N 标志位自动转换为 0/1
        if self._is_yn_flag(fm):
            return (f"CASE WHEN {alias}.{fm.source_field} = 'Y' "
                    f"THEN '1' ELSE '0' END")

        # 9. 检测可能需要字典码值转换的字段 — 优先从字典查找
        dict_expr = self._lookup_case_dict(fm, seg)
        if dict_expr:
            return dict_expr
        self._check_dict_mismatch(fm)

        # 10. 普通直取
        return f"{alias}.{fm.source_field}"

    @staticmethod
    def _is_function_expr(field: str) -> bool:
        """判断字段是否包含函数调用。
        匹配两种模式：
        1. 直接函数: NVL(...), SUM(...)
        2. 别名前缀函数: T1.SUM(...), A.MAX(...)
        """
        f = field.strip()
        funcs = (
            r'NVL|IFNULL|COALESCE|CONCAT|SUBSTR|TRIM'
            r'|SUM|MAX|MIN|COUNT|AVG'
            r'|WM_CONCAT|GROUP_CONCAT'
        )
        # 模式1: FUNC(...)
        if re.match(rf'^({funcs})\s*\(', f, re.IGNORECASE):
            return True
        # 模式2: ALIAS.FUNC(...) — 如 T1.SUM(...), A.MAX(...)
        if re.match(
            rf'^[A-Z]\w*\.({funcs})\s*\(', f, re.IGNORECASE
        ):
            return True
        return False

    def _convert_source_field_expr(
        self, fm: FieldMapping, seg: MappingSegment
    ) -> str:
        """将源字段中的函数表达式转为 MySQL 语法"""
        expr = fm.source_field.strip()

        # 清洗全角字符
        expr = ExcelParser._clean_sql_text(expr)

        # 替换旧别名 (A.X → T1.X)
        if seg.alias_map:
            expr = ExcelParser._replace_aliases(
                expr, seg.alias_map)

        # 修复 ALIAS.FUNC(...) → FUNC(ALIAS....)
        # 如 T1.SUM(AMT) → SUM(T1.AMT)
        m = re.match(
            r'^(\w+)\.(SUM|MAX|MIN|COUNT|AVG|WM_CONCAT'
            r'|GROUP_CONCAT)\s*\((.+)\)\s*$',
            expr, re.IGNORECASE
        )
        if m:
            alias_part, func_name, args = (
                m.group(1), m.group(2), m.group(3))
            if '.' not in args:
                args = f'{alias_part}.{args}'
            func_upper = func_name.upper()
            if func_upper == 'WM_CONCAT':
                func_upper = 'GROUP_CONCAT'
            return f'{func_upper}({args})'

        # 直接聚合函数：给裸列名添加别名前缀
        # 如 SUM(MWK_AMT) → SUM(T1.MWK_AMT)
        alias = self._resolve_alias(fm, seg)
        m = re.match(
            r'^(SUM|MAX|MIN|COUNT|AVG|GROUP_CONCAT'
            r'|WM_CONCAT)\s*\((.+)\)\s*$',
            expr, re.IGNORECASE
        )
        if m:
            func_name, args = m.group(1), m.group(2)
            func_upper = func_name.upper()
            if func_upper == 'WM_CONCAT':
                func_upper = 'GROUP_CONCAT'
            # 给裸列名添加别名
            args = re.sub(
                r'\b([A-Z_]\w+)\b',
                lambda mm: (
                    mm.group(0)
                    if '.' in mm.group(0)
                    or mm.group(0).upper() in (
                        'SEPARATOR', 'DISTINCT',
                        'ASC', 'DESC')
                    else f'{alias}.{mm.group(0)}'
                ),
                args
            )
            return f'{func_upper}({args})'

        # NVL(...) → IFNULL(...)
        m = re.match(
            r'NVL\s*\((.+?),\s*(.+?)\)\s*$',
            expr, re.IGNORECASE
        )
        if m:
            return f"IFNULL({m.group(1)}, {m.group(2)})"

        # WM_CONCAT → GROUP_CONCAT
        m = re.match(
            r'WM_CONCAT\s*\((.+)\)\s*$',
            expr, re.IGNORECASE
        )
        if m:
            return f"GROUP_CONCAT({m.group(1)})"

        # COALESCE 直接返回
        if expr.upper().startswith('COALESCE'):
            return expr

        # 其他函数直接返回
        self._note(
            f"字段 {fm.target_en_name}({fm.target_cn_name}) "
            f"源字段包含函数表达式: '{expr}'"
        )
        return expr

    def _gen_conditional_fill(
        self, fm: FieldMapping, seg: MappingSegment
    ) -> str:
        """根据填报说明中的条件逻辑生成 CASE WHEN 表达式。
        如 '当渠道类型为ATM机/VTM/POS时，填交易渠道号'
        """
        inst = fm.fill_instruction
        alias = self._resolve_alias(fm, seg)

        # 模式: "当渠道类型为XX时，填YYY"
        # 或: "当...为柜面时，填..."
        m = re.search(r'当.*?为(.+?)时', inst)
        if not m:
            return ''

        condition_text = m.group(1).strip()

        # 查找引用的条件字段 - 常见模式：渠道类型
        cond_field = ''
        if '渠道' in inst:
            cond_field = f'{alias}.CHANNEL_TYPE_CODE'

        if not cond_field:
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name}) "
                f"填报说明有条件逻辑但无法自动识别条件字段: "
                f"'{inst}'"
            )
            return ''

        # 解析条件值
        codes = []
        # 柜面 → 01
        channel_map = {
            '柜面': '01', 'ATM': '02', 'ATM机': '02',
            '自助终端': '02', 'VTM': '03', 'POS': '04',
            '网银': '05', '手机银行': '06', '手机': '06',
            '第三方支付': '07', '银联': '08',
        }
        # 按 \\ / 、 分隔条件文本
        parts = re.split(r'[\\\\、/]', condition_text)
        for p in parts:
            p = p.strip()
            if p in channel_map:
                codes.append(channel_map[p])

        if not codes:
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name}) "
                f"无法解析条件值: '{condition_text}'"
            )
            return ''

        # 去重并保持顺序
        seen = set()
        codes = [c for c in codes if not (c in seen or seen.add(c))]

        if len(codes) == 1:
            cond = f"{cond_field} = '{codes[0]}'"
        else:
            in_list = ', '.join(f"'{c}'" for c in codes)
            cond = f"{cond_field} IN ({in_list})"

        return (
            f"CASE WHEN {cond} "
            f"THEN {alias}.{fm.source_field} ELSE '' END"
        )

    def _convert_mapping_rule(
        self, fm: FieldMapping, seg: MappingSegment
    ) -> str:
        """将 Excel 中的映射规则转为 MySQL 表达式"""
        rule = fm.mapping_rule

        # 清洗全角字符
        rule = ExcelParser._clean_sql_text(rule)
        # 通用 Oracle 语法转换
        rule = ExcelParser._convert_oracle_syntax(rule)

        # Oracle TO_CHAR → MySQL DATE_FORMAT
        match = re.match(
            r"DATE_FORMAT\((\w+\.\w+)\s*,\s*'%Y-%m-%d'\)",
            rule, re.IGNORECASE)
        if match:
            field_ref = match.group(1)
            default = self._get_date_default(fm)
            expr = f"DATE_FORMAT({field_ref}, '%Y-%m-%d')"
            if default:
                expr = f"COALESCE({expr}, '{default}')"
            return expr

        # IF ... THEN ... ELSE ... END IF → CASE WHEN
        match = re.match(
            r"IF\s+(.+?)\s+THEN\s*\n?\s*'(.+?)'\s*\n?\s*ELSE\s*\n?\s*'(.+?)'\s*\n?\s*END\s*IF",
            rule, re.IGNORECASE | re.DOTALL
        )
        if match:
            condition = match.group(1).strip()
            then_val = match.group(2)
            else_val = match.group(3)
            # 转换 Or → OR
            condition = re.sub(r'\bOr\b', 'OR', condition)
            # 给裸列名添加表别名
            alias = self._resolve_alias(fm, seg)
            condition = self._qualify_columns(
                condition, alias
            )
            return (f"CASE WHEN {condition} "
                    f"THEN '{then_val}' ELSE '{else_val}' END")

        # NVL → IFNULL
        match = re.match(r"NVL\((.+?),\s*(.+?)\)", rule, re.IGNORECASE)
        if match:
            return f"IFNULL({match.group(1)}, {match.group(2)})"

        # 通过产品类别进行区分 → 占位
        if '产品类别' in rule or '区分' in rule:
            self._note(f"字段 {fm.target_en_name}({fm.target_cn_name}) 映射规则 '{rule}' 需人工补充具体逻辑")
            return "'0'"

        # 需转换标记
        if rule.strip() in ('转换', '需转换'):
            # 优先从字典查找 CASE 映射
            dict_expr = self._lookup_case_dict(fm, seg)
            if dict_expr:
                return dict_expr
            self._note(f"字段 {fm.target_en_name}({fm.target_cn_name}) 标注'需转换'，源字典值未知，暂直取")
            alias = self._resolve_alias(fm, seg)
            return f"{alias}.{fm.source_field}"

        # CASE WHEN ... END 表达式 — 已是合法 SQL，直接使用
        if re.match(r'CASE\b', rule, re.IGNORECASE) and re.search(r'\bEND\b', rule, re.IGNORECASE):
            # 修正 Oracle IS 'val' → = 'val'
            expr = re.sub(r"\bIS\s+'", "= '", rule, flags=re.IGNORECASE)
            # 给裸列名添加表别名
            alias = self._resolve_alias(fm, seg)
            expr = self._qualify_columns(expr, alias)
            # 替换别名映射
            if seg.alias_map:
                expr = ExcelParser._replace_aliases(expr, seg.alias_map)
            return expr

        # 无法识别的规则，直接输出为注释
        self._note(f"字段 {fm.target_en_name}({fm.target_cn_name}) 映射规则无法自动转换: '{rule}'")
        alias = self._resolve_alias(fm, seg)
        return f"{alias}.{fm.source_field}  /* TODO: {rule} */"

    def _gen_multi_source_expr(self, fm: FieldMapping, extra_field: str, seg: MappingSegment) -> str:
        """处理多源字段（如贷款逾期标识：OVERDUE_BAL 和 INT_OVERDUE_BAL）"""
        alias = self._resolve_alias(fm, seg)
        # 检查映射规则中是否有 IF 逻辑已嵌入
        # 默认生成 CASE WHEN 逻辑
        return f"CASE WHEN {alias}.{fm.source_field} > 0 OR {alias}.{extra_field} > 0 THEN '1' ELSE '0' END"

    def _resolve_alias(self, fm: FieldMapping, seg: MappingSegment) -> str:
        """根据源表名解析表别名"""
        src = fm.source_table
        if not src:
            return seg.source_tables[0].alias if seg.source_tables else 'T1'

        # 如果 source_table 已经是别名（如 T3, T4）
        if re.match(r'^T\d+$', src):
            return src

        # 按表名匹配
        for t in seg.source_tables:
            if t.table_name == src:
                return t.alias

        # 模糊匹配
        for t in seg.source_tables:
            if src in t.table_name or t.table_name in src:
                return t.alias

        return seg.source_tables[0].alias if seg.source_tables else 'T1'

    @staticmethod
    def _qualify_columns(expr: str, alias: str) -> str:
        """给 SQL 表达式中未限定表别名的列名添加别名前缀。
        如 SPEC_ACCT_TYPE_CODE='103' → T1.SPEC_ACCT_TYPE_CODE='103'
        """
        def replacer(m):
            col = m.group(0)
            # 排除 SQL 关键字和常见函数
            keywords = {
                'AND', 'OR', 'NOT', 'IN', 'IS', 'NULL',
                'LIKE', 'BETWEEN', 'CASE', 'WHEN', 'THEN',
                'ELSE', 'END', 'IF', 'TRUE', 'FALSE',
                'DATE_FORMAT', 'COALESCE', 'IFNULL', 'NVL',
                'CONCAT', 'SUBSTR', 'TRIM', 'UPPER', 'LOWER',
            }
            if col.upper() in keywords:
                return col
            return f'{alias}.{col}'
        # 匹配未被 . 前缀的标识符（后跟 = < > 空格 等）
        return re.sub(
            r'(?<![.\w])([A-Z_][A-Z0-9_]{2,})(?=\s*[=<>!])',
            lambda m: replacer(m), expr
        )

    @staticmethod
    def _is_yn_flag(fm: FieldMapping) -> bool:
        """判断字段是否为 Y/N 标志位且需要转为 0/1。
        识别信号：源字段名含 _FLAG，且满足以下任一条件：
        1. 目标字典含 0/1 或 是/否
        2. 目标类型为短 VARCHAR（1-2位），暗示码值字段
        """
        if not fm.source_field:
            return False
        # 源字段名含 _FLAG
        if '_FLAG' not in fm.source_field.upper():
            return False
        # 条件1: 目标字典暗示 0/1
        td = fm.target_dict or ''
        if re.search(r'[01].*[否是]|[是否].*[01]|0[-.:]否|1[-.:]是', td):
            return True
        if '0' in td and '1' in td:
            return True
        # 条件2: 目标类型为 VARCHAR(1) 或 VARCHAR(2)，字典列为空
        # _FLAG 字段 + 短 VARCHAR 目标 = 几乎必然是 Y/N→0/1
        tgt_type = (fm.target_type or '').upper()
        m = re.search(r'VARCHAR\w*\((\d+)\)', tgt_type)
        if m and int(m.group(1)) <= 2 and not td:
            return True
        return False

    def _check_dict_mismatch(self, fm: FieldMapping):
        """检测可能需要字典码值转换的字段，生成警告"""
        src_type = fm.source_type.upper() if fm.source_type else ''
        tgt_type = fm.target_type.upper() if fm.target_type else ''

        # 提取 VARCHAR 宽度: VARCHAR2(6) → 6
        src_w = re.search(r'VARCHAR\w*\((\d+)\)', src_type)
        tgt_w = re.search(r'VARCHAR\w*\((\d+)\)', tgt_type)

        if src_w and tgt_w:
            sw, tw = int(src_w.group(1)), int(tgt_w.group(1))
            if sw > tw and tw <= 2:
                # 源类型宽度 > 目标类型宽度，且目标是短码
                desc = fm.description or ''
                if re.search(r'\d{2}\s+\S', desc):
                    self._note(
                        f"字段 {fm.target_en_name}"
                        f"({fm.target_cn_name}) "
                        f"源类型 {fm.source_type} → "
                        f"目标类型 {fm.target_type}，"
                        f"业务口径含码值定义，"
                        f"可能需要 CASE 转换"
                    )

        # 源为 INTEGER 目标为 VARCHAR(2) — 可能需要格式化
        if 'INTEGER' in src_type or 'INT' in src_type:
            if 'VARCHAR' in tgt_type and tgt_w:
                if int(tgt_w.group(1)) <= 2:
                    self._note(
                        f"字段 {fm.target_en_name}"
                        f"({fm.target_cn_name}) "
                        f"源类型 {fm.source_type} → "
                        f"目标类型 {fm.target_type}，"
                        f"可能需要数值→字符串转换"
                    )

    def _lookup_case_dict(
        self, fm: FieldMapping, seg: MappingSegment
    ) -> Optional[str]:
        """从 CASE 字典中查找映射。命中返回 CASE 表达式，未命中返回 None。"""
        if not self.case_dict:
            return None
        source_field = fm.source_field
        if not source_field:
            return None
        # 去掉聚合函数包装：MAX(FIELD) → FIELD
        m = re.match(r'\w+\((\w+)\)', source_field)
        if m:
            source_field = m.group(1)
        # 去掉全角括号清洗后的残留
        source_field = re.sub(r'[（）]', '', source_field).strip()

        expr = self.case_dict.lookup(fm.target_en_name, source_field)
        if expr:
            # 替换表别名为当前段的实际别名
            alias = self._resolve_alias(fm, seg)
            # 将手写 SQL 中的 Tn. 前缀统一替换为当前段的别名
            result = re.sub(r'\bT\d+\.', f'{alias}.', expr)
            # 清理空行
            result = re.sub(r'\n\s*\n', '\n', result)
            self._note(
                f"字段 {fm.target_en_name}({fm.target_cn_name}) "
                f"使用字典中的 CASE 映射（源: {source_field}）"
            )
            return result
        return None

    def _is_date_conversion(self, fm: FieldMapping) -> bool:
        """判断是否需要日期格式转换"""
        src_type = fm.source_type.upper() if fm.source_type else ''
        tgt_type = fm.target_type.upper() if fm.target_type else ''
        return ('DATE' in src_type and 'DATE' not in tgt_type and 'VARCHAR' in tgt_type)

    def _get_date_default(self, fm: FieldMapping) -> str:
        """获取日期字段的默认值"""
        desc = ((fm.fill_instruction or '') +
                (fm.description or '') +
                (fm.target_cn_name or ''))
        if '9999-12-31' in desc:
            return '9999-12-31'
        if '9999-12' in desc:
            return '9999-12'
        return ''

    def _gen_footer(self) -> str:
        """生成存储过程尾部"""
        return """    SET O_RLT = 'true';

    -- 记录日志
    CALL dwdevdb_model.PMODEL_JOB_LOG(V_DATE, V_PRD_NAME, V_TAB_NAME, V_TAGS,
                      O_RLT, V_TOTAL_NUM, V_START_DT, NOW());

END;
"""


# ============================================================
# 主程序
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='根据 Excel 映射定义生成 MySQL 存储过程 SQL')
    parser.add_argument('excel_file', help='Excel 文件路径')
    parser.add_argument('sheet_name', help='Sheet 名称')
    parser.add_argument('-o', '--output', help='输出 SQL 文件路径（默认输出到标准输出）')
    parser.add_argument('-v', '--verbose', action='store_true', help='显示详细信息')
    parser.add_argument('--dict-from', dest='dict_from', metavar='DIR',
                        action='append', default=[],
                        help='从指定目录的手写 SQL 文件中提取 CASE 映射字典（可多次指定）')
    args = parser.parse_args()

    # 加载 CASE 字典
    case_dict = None
    if args.dict_from:
        case_dict = CaseDictExtractor()
        total = 0
        for d in args.dict_from:
            n = case_dict.load_from_directory(d)
            total += n
            print(f"[信息] 从 {d} 加载了 {n} 个 CASE 映射", file=sys.stderr)
        if len(args.dict_from) > 1:
            print(f"[信息] 合计 {total} 个 CASE 映射", file=sys.stderr)

    # 解析 Excel
    ep = ExcelParser(args.excel_file, args.sheet_name)
    mapping = ep.parse()

    # 输出警告和错误
    for w in ep.warnings:
        print(w, file=sys.stderr)
    for e in ep.errors:
        print(e, file=sys.stderr)

    if not mapping:
        print("\n解析失败，请检查以上错误信息。", file=sys.stderr)
        sys.exit(1)

    if args.verbose:
        print(f"\n目标表: {mapping.target_table} ({mapping.target_cn_name})", file=sys.stderr)
        for seg in mapping.segments:
            print(f"  段: {seg.segment_name}", file=sys.stderr)
            print(f"    源表: {[t.table_name + ' ' + t.alias for t in seg.source_tables]}", file=sys.stderr)
            print(f"    条件: {len(seg.where_conditions)} 个", file=sys.stderr)
            print(f"    字段: {len(seg.field_mappings)} 个", file=sys.stderr)

    # 生成 SQL
    gen = SQLGenerator(mapping, case_dict=case_dict)
    sql = gen.generate()

    # 输出注意事项
    for n in gen.notes:
        print(n, file=sys.stderr)

    # 输出 SQL
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(sql)
        print(f"\nSQL 已写入: {args.output}", file=sys.stderr)
    else:
        print(sql)


if __name__ == '__main__':
    main()
