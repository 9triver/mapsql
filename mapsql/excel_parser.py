"""Excel sheet parser: reads Excel and produces SheetMapping models."""

import re
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

from .models import (
    SourceTable, WhereCondition, FieldMapping, MappingSegment, SheetMapping,
)
from . import text_cleaner


class ExcelParser:
    """Parse an Excel Sheet's mapping definition into a SheetMapping model."""

    def __init__(self, filepath: str, sheet_name: str):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.warnings: list[str] = []
        self.errors: list[str] = []

    def _warn(self, row: int, msg: str):
        self.warnings.append(f"[警告] 行{row}: {msg}")

    def _error(self, row: int, msg: str):
        self.errors.append(f"[错误] 行{row}: {msg}")

    def _cell(self, row, col) -> str:
        val = self.ws.cell(row=row, column=col).value
        if val is None:
            return ''
        return str(val).strip()

    def parse(self) -> Optional[SheetMapping]:
        if openpyxl is None:
            self.errors.append("[错误] 需要 openpyxl 库。请运行: pip install openpyxl")
            return None

        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
        except FileNotFoundError:
            self.errors.append(f"[错误] 文件不存在: {self.filepath}")
            return None
        except Exception as e:
            self.errors.append(f"[错误] 无法打开文件: {e}")
            return None

        if self.sheet_name not in wb.sheetnames:
            self.errors.append(
                f"[错误] Sheet '{self.sheet_name}' 不存在。"
                f"可用的 Sheet: {', '.join(wb.sheetnames)}"
            )
            return None

        self.ws = wb[self.sheet_name]
        max_row = self.ws.max_row

        target_table, target_cn_name = self._parse_target_table()
        if not target_table:
            return None

        mapping = SheetMapping(
            target_table=target_table,
            target_cn_name=target_cn_name,
        )

        segments_raw = self._split_segments(max_row)
        for seg_name, seg_rows in segments_raw:
            segment = self._parse_segment(seg_name, seg_rows)
            if segment:
                mapping.segments.append(segment)

        if not mapping.segments:
            self.errors.append("[错误] 未找到任何有效的映射段")
            return None

        return mapping

    def _parse_target_table(self) -> tuple[str, str]:
        for r in range(1, min(self.ws.max_row + 1, 5)):
            if self._cell(r, 1) == '目标表：':
                table_name = self._cell(r, 2)
                table_cn = self._cell(r, 3)
                if not table_name:
                    self._error(r, "目标表名为空")
                    return '', ''
                return table_name, table_cn
        self._error(0, "未找到 '目标表：' 定义行。预期格式：第1列为'目标表：'，第2列为表英文名，第3列为表中文名")
        return '', ''

    def _split_segments(self, max_row: int) -> list:
        segment_starts = []
        seg_name = ''
        for r in range(1, max_row + 1):
            c1 = self._cell(r, 1)
            if re.match(r'^[A-Z]段[：:]', c1):
                seg_name = c1
                continue
            if c1 == '数据源表：':
                name = seg_name if seg_name else '默认段'
                segment_starts.append((name, r))
                seg_name = ''

        if not segment_starts:
            for r in range(1, max_row + 1):
                if self._cell(r, 1) == '数据源表：':
                    segment_starts.append(('默认段', r))
                    break

        if not segment_starts:
            self._error(0, "未找到 '数据源表：' 定义行")
            return []

        result = []
        for i, (name, start_row) in enumerate(segment_starts):
            if i + 1 < len(segment_starts):
                end_row = segment_starts[i + 1][1] - 1
            else:
                end_row = max_row
            result.append((name, (start_row, end_row)))

        return result

    def _parse_segment(self, seg_name: str, row_range: tuple) -> Optional[MappingSegment]:
        start_row, end_row = row_range
        segment = MappingSegment(segment_name=seg_name)

        source_table_start = start_row
        where_start = None
        field_start = None

        for r in range(start_row, end_row + 1):
            c1 = self._cell(r, 1)
            if c1 == '数据范围条件：':
                where_start = r
            if c1 == '字段映射':
                field_start = r

        # Parse source tables
        source_end = (where_start - 1) if where_start else (
            field_start - 1 if field_start else end_row)
        segment.source_tables = self._parse_source_tables(
            source_table_start, source_end)
        segment.alias_map = getattr(self, '_last_alias_map', {})

        if not segment.source_tables:
            self._error(start_row, f"段 '{seg_name}' 未找到任何数据源表定义")
            return None

        # Parse WHERE conditions
        if where_start:
            where_end = (field_start - 1) if field_start else end_row
            segment.where_conditions = self._parse_where_conditions(where_start, where_end)
            # Apply alias mapping to WHERE conditions
            if segment.alias_map:
                for wc in segment.where_conditions:
                    wc.condition = text_cleaner.replace_aliases(
                        wc.condition, segment.alias_map)
            # Expand "无" join type tables in WHERE subqueries
            for t in segment.source_tables:
                if t.join_type == '' and t != segment.source_tables[0]:
                    wc_alias = t.alias
                    for wc in segment.where_conditions:
                        wc.condition = re.sub(
                            rf'\bFrom\s+{re.escape(wc_alias)}\b',
                            f'From {t.table_name} {wc_alias}',
                            wc.condition,
                            flags=re.IGNORECASE,
                        )

        # Parse field mappings
        if field_start:
            segment.field_mappings = self._parse_field_mappings(field_start, end_row)
        else:
            self._error(start_row, f"段 '{seg_name}' 未找到 '字段映射' 标记行")
            return None

        if not segment.field_mappings:
            self._error(start_row, f"段 '{seg_name}' 未解析到任何字段映射")
            return None

        return segment

    def _parse_source_tables(self, start_row: int, end_row: int) -> list:
        tables = []
        for r in range(start_row + 1, end_row + 1):
            table_name = self._cell(r, 2)
            if not table_name:
                continue
            table_cn = self._cell(r, 3)
            alias_raw = self._cell(r, 4)
            join_type_raw = self._cell(r, 5)
            join_cond_raw = self._cell(r, 6)
            remark = self._cell(r, 8)

            alias = text_cleaner.extract_alias(alias_raw, len(tables) + 1)
            if not alias:
                self._warn(r, f"表 {table_name} 缺少别名，自动分配 T{len(tables) + 1}")
                alias = f'T{len(tables) + 1}'

            join_type = ''
            if join_type_raw:
                jt = join_type_raw.upper().strip()
                if jt == '无' or jt == 'NONE':
                    join_type = ''
                elif 'INNER' in jt:
                    join_type = 'INNER JOIN'
                elif 'LEFT' in jt:
                    join_type = 'LEFT JOIN'
                elif 'RIGHT' in jt:
                    join_type = 'RIGHT JOIN'
                else:
                    self._warn(r, f"无法识别关联类型 '{join_type_raw}'，默认为 LEFT JOIN")
                    join_type = 'LEFT JOIN'

            join_cond = text_cleaner.clean_sql_text(join_cond_raw)
            if join_cond.upper().startswith('ON '):
                join_cond = join_cond[3:].strip()

            tables.append(SourceTable(
                table_name=table_name,
                table_cn_name=table_cn,
                alias=alias,
                join_type=join_type,
                join_condition=join_cond,
                remark=remark,
            ))

        if tables and tables[0].join_type:
            self._warn(start_row + 1, f"第一个表 {tables[0].table_name} 不应有关联类型（应为主表）")

        # Unify aliases to Tn format
        alias_map = {}
        for i, t in enumerate(tables):
            new_alias = f'T{i + 1}'
            if t.alias != new_alias and re.match(r'^[A-Z]$', t.alias):
                alias_map[t.alias] = new_alias
                t.alias = new_alias

        if alias_map:
            for t in tables:
                t.join_condition = text_cleaner.replace_aliases(
                    t.join_condition, alias_map
                )

        self._last_alias_map = alias_map
        return tables

    def _parse_where_conditions(self, start_row: int, end_row: int) -> list:
        conditions = []
        for r in range(start_row + 1, end_row + 1):
            operator = self._cell(r, 2).upper()
            condition = text_cleaner.clean_sql_text(self._cell(r, 3))
            description = self._cell(r, 5)
            if not operator or not condition:
                continue
            # Convert Chinese pseudo-function: 月初(X) → DATE_FORMAT(X, '%Y-%m-01')
            condition = re.sub(
                r'月初\((\w+)\)',
                r"DATE_FORMAT(\1, '%Y-%m-01')",
                condition,
            )
            condition = text_cleaner.convert_oracle_syntax(condition)
            conditions.append(WhereCondition(
                operator=operator,
                condition=condition,
                description=description,
            ))
        return conditions

    def _parse_field_mappings(self, start_row: int, end_row: int) -> list:
        mappings = []

        header_row = None
        for r in range(start_row, min(start_row + 5, end_row + 1)):
            if self._cell(r, 1) == '字段中文名':
                header_row = r
                break

        if header_row is None:
            for r in range(start_row, min(start_row + 5, end_row + 1)):
                if '目标字段' in self._cell(r, 1) or '字段中文名' in self._cell(r, 1):
                    header_row = r
                    break

        if header_row is None:
            self._error(start_row, "字段映射区未找到标题行（'字段中文名'）")
            return []

        for r in range(header_row + 1, end_row + 1):
            cn_name = self._cell(r, 1)
            en_name = self._cell(r, 2)

            if not cn_name and not en_name:
                continue

            # Continuation row: extra source field for previous mapping
            if not en_name and not self._cell(r, 3):
                src_field = self._cell(r, 7)
                if src_field and mappings:
                    mappings[-1].mapping_rule = mappings[-1].mapping_rule or ''
                    if not mappings[-1].mapping_rule:
                        mappings[-1].mapping_rule = f'__MULTI_SRC__:{src_field}'
                continue

            target_type = self._cell(r, 3)
            target_dict = self._cell(r, 4)
            source_table = self._cell(r, 6)
            source_field = self._cell(r, 7)
            source_cn = self._cell(r, 8)
            source_type = self._cell(r, 9)
            mapping_rule = self._cell(r, 10)
            source_dict = self._cell(r, 11)
            fill_instruction = self._cell(r, 12)
            description = self._cell(r, 13)
            biz_scope = self._cell(r, 14)

            if en_name and not target_type and cn_name != '采集日期':
                self._warn(r, f"字段 {en_name}({cn_name}) 缺少目标字段类型")

            mappings.append(FieldMapping(
                target_cn_name=cn_name,
                target_en_name=en_name,
                target_type=target_type,
                target_dict=target_dict,
                source_table=source_table,
                source_field=source_field,
                source_cn_name=source_cn,
                source_type=source_type,
                mapping_rule=mapping_rule,
                source_dict=source_dict,
                fill_instruction=fill_instruction,
                description=description,
                biz_scope=biz_scope,
            ))

        return mappings
